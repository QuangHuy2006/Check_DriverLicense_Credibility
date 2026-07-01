import os
import json
import asyncio
import time
import base64
from datetime import datetime, timedelta
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, UploadFile, File, BackgroundTasks
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from collections import defaultdict
import httpx
from bs4 import BeautifulSoup
import re
import io
import openpyxl
import logging
from logging.handlers import RotatingFileHandler
from contextlib import asynccontextmanager

# === LOGGING SETUP ===
LOG_DIR = os.path.dirname(os.path.abspath(__file__))
log_handler = RotatingFileHandler(
    os.path.join(LOG_DIR, "activity.log"),
    maxBytes=5*1024*1024,  # 5MB
    backupCount=3,
    encoding="utf-8"
)
log_handler.setFormatter(logging.Formatter(
    "%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S"
))
logger = logging.getLogger("gplx")
logger.setLevel(logging.INFO)
logger.addHandler(log_handler)

import database as db
from app_ocr import doc_captcha_ddddocr, ocr

# --- CLEANUP TASK: Dọn ip_requests cũ mỗi 5 phút để tránh memory leak ---
async def _cleanup_ip_requests():
    while True:
        await asyncio.sleep(300)  # 5 phút
        now = datetime.now()
        for ip in list(ip_requests.keys()):
            ip_requests[ip] = [t for t in ip_requests[ip] if (now - t).seconds < 60]
            if not ip_requests[ip]:
                del ip_requests[ip]

@asynccontextmanager
async def lifespan(app):
    asyncio.create_task(_cleanup_ip_requests())
    yield

app = FastAPI(title="GPLX Verification System", lifespan=lifespan)

# Cấu hình tĩnh
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
static_dir = os.path.join(BASE_DIR, "static")
os.makedirs(static_dir, exist_ok=True)
app.mount("/static", StaticFiles(directory=static_dir), name="static")

# --- SEMAPHORE: Giới hạn số request đồng thời tới CSGT ---
# Tối đa 2 request cùng lúc, những người còn lại xếp hàng chờ
CSGT_SEMAPHORE = asyncio.Semaphore(2)

# --- PRELOAD STATE: Theo dõi tiến trình pre-load ban đêm ---
preload_state = {
    "running": False,
    "stop_requested": False,
    "total": 0,
    "done": 0,
    "skipped": 0,   # đã có trong DB
    "failed": 0,
    "current": "",
    "errors": [],
    "started_at": None,
}

# --- RATE LIMITER ---
RATE_LIMIT_MINUTES = 1
MAX_REQUESTS_PER_MINUTE = 200  # Nới rộng cho tra cứu hàng loạt
ip_requests = defaultdict(list)

def check_rate_limit(client_ip: str) -> tuple[bool, str]:
    now = datetime.now()
    ip_requests[client_ip] = [req_time for req_time in ip_requests[client_ip] if now - req_time < timedelta(minutes=RATE_LIMIT_MINUTES)]
    
    if len(ip_requests[client_ip]) >= MAX_REQUESTS_PER_MINUTE:
        return False, f"Bạn đã vượt quá giới hạn {MAX_REQUESTS_PER_MINUTE} yêu cầu/phút. Vui lòng thử lại sau."
    
    ip_requests[client_ip].append(now)
    return True, ""

# --- API ENDPOINTS ---

@app.get("/")
async def get_index():
    return FileResponse(os.path.join(static_dir, "index.html"))

# Mật khẩu quản trị (bạn có thể thay đổi tại đây)
ADMIN_SECRET = os.environ.get("ADMIN_SECRET", "@@@@@@@@")

@app.get("/api/export-excel")
async def export_excel(key: str = ""):
    if key != ADMIN_SECRET:
        logger.warning(f"Export bị từ chối - sai mật khẩu")
        return JSONResponse({"error": "Không có quyền truy cập."}, status_code=403)
    logger.info(f"Admin xuất file Excel")
    # Fix #4: db call là synchronous, chạy trong thread pool
    history = await asyncio.to_thread(db.get_all_history)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lich Su GPLX"
    
    ws.append(["Họ Tên", "Số GPLX", "Loại Bằng", "Ngày Cấp", "Thời Hạn", "Trạng Thái"])
    
    for row in history:
        name = row.get("name", "")
        gplx = row.get("gplx", "")
        loai_bang = row.get("loai_bang", "")
        ngay_cap = row.get("ngay_cap", "")
        thoi_han = row.get("thoi_han", "")
        status_map = {"success": "Thành công", "not_found": "Không tìm thấy", "error": "Lỗi"}
        trang_thai = status_map.get(row.get("status", ""), row.get("status", ""))
        
        ws.append([name, gplx, loai_bang, ngay_cap, thoi_han, trang_thai])
    
    export_path = os.path.join(BASE_DIR, "Lich_su_GPLX.xlsx")
    wb.save(export_path)
    
    return FileResponse(
        export_path, 
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename="Lich_su_GPLX.xlsx"
    )

# ===========================
# PRE-LOAD ENDPOINTS (Admin)
# ===========================

@app.post("/api/admin/start-preload")
async def start_preload(background_tasks: BackgroundTasks, file: UploadFile = File(...), key: str = "", delay: float = 3.0):
    """Upload Excel và bắt đầu pre-load toàn bộ danh sách vào DB ở background."""
    if key != ADMIN_SECRET:
        return JSONResponse({"error": "Không có quyền truy cập."}, status_code=403)
    if preload_state["running"]:
        return JSONResponse({"error": "Đang có tiến trình pre-load đang chạy. Dừng lại trước."}, status_code=400)
    
    contents = await file.read()
    wb = openpyxl.load_workbook(filename=io.BytesIO(contents))
    ws = wb.active
    
    records = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue
        if not row[0] or not row[1]:
            continue
        gplx = str(row[0]).strip()
        raw_dob = row[1]
        if isinstance(raw_dob, datetime):
            dob = raw_dob.strftime("%d/%m/%Y")
        else:
            dob = str(raw_dob).strip()
        loai = str(row[2]).strip().upper() if len(row) > 2 and row[2] else "PET"
        records.append({
            "gplx": gplx,
            "dob": dob,
            "loai_bang": "PET" if "PET" in loai or "2" in loai else "OLD"
        })
    
    if not records:
        return JSONResponse({"error": "Không tìm thấy dữ liệu hợp lệ trong file."}, status_code=400)
    
    # Reset state
    preload_state.update({
        "running": True,
        "stop_requested": False,
        "total": len(records),
        "done": 0,
        "skipped": 0,
        "failed": 0,
        "current": "",
        "errors": [],       # giới hạn 100 dòng lỗi gần nhất
        "started_at": datetime.now().strftime("%H:%M:%S %d/%m/%Y"),
    })
    
    background_tasks.add_task(_run_preload, records, delay)
    logger.info(f"Pre-load bắt đầu: {len(records)} bản ghi, delay={delay}s")
    return JSONResponse({"status": "started", "total": len(records)})

@app.post("/api/admin/stop-preload")
async def stop_preload(key: str = ""):
    if key != ADMIN_SECRET:
        return JSONResponse({"error": "Không có quyền truy cập."}, status_code=403)
    preload_state["stop_requested"] = True
    return JSONResponse({"status": "stop_requested"})

@app.get("/api/admin/preload-status")
async def get_preload_status(key: str = ""):
    if key != ADMIN_SECRET:
        return JSONResponse({"error": "Không có quyền truy cập."}, status_code=403)
    return JSONResponse(preload_state)

async def _run_preload(records: list, delay: float):
    """Background task: tuần tự tra cứu từng GPLX và lưu vào DB."""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7",
    }

    for item in records:
        if preload_state["stop_requested"]:
            logger.info("Pre-load bị dừng theo yêu cầu.")
            break

        gplx = item["gplx"]
        dob = item["dob"]
        loai_bang = item["loai_bang"]
        preload_state["current"] = gplx

        # Nếu đã có trong DB → bỏ qua
        already = await asyncio.to_thread(db.is_gplx_verified, gplx)
        if already:
            stored_dob = await asyncio.to_thread(db.get_stored_dob, gplx)
            if stored_dob == dob:
                preload_state["skipped"] += 1
                logger.info(f"[Pre-load] Bỏ qua (đã có DB): {gplx}")
                continue

        try:
            async with CSGT_SEMAPHORE:
                async with httpx.AsyncClient(verify=False, timeout=20.0) as client:
                    resp = await client.get("https://gplx.csgt.bocongan.gov.vn/", headers=headers)
                soup = BeautifulSoup(resp.text, 'html.parser')
                sec_token_input = soup.find('input', {'name': 'securityToken'})
                if not sec_token_input:
                    raise Exception("Không lấy được securityToken")
                security_token = sec_token_input.get('value')
                choose_gplx = "1" if loai_bang == "OLD" else "2"

                # Lấy captcha
                captcha_img_tag = soup.select_one(".img-cap-mobile img")
                if not captcha_img_tag:
                    raise Exception("Không tìm thấy ảnh captcha")
                captcha_url = "https://gplx.csgt.bocongan.gov.vn" + captcha_img_tag.get('src')
                if '?' in captcha_url:
                    captcha_url = re.sub(r't=\d+', f't={int(time.time()*1000)}', captcha_url)

                success_result = None
                for attempt in range(3):
                    if preload_state["stop_requested"]:
                        break
                    cap_resp = await client.get(captcha_url, headers=headers)
                    image_bytes = cap_resp.content
                    cap_code = await asyncio.to_thread(ocr.classification, image_bytes)

                    if not cap_code or not (4 <= len(cap_code) <= 8):
                        continue

                    payload = {
                        "type": "",
                        "fields[formTypeId]": "565f96637f8b9af6558b4567",
                        "fields[chooseGPLX]": choose_gplx,
                        "fields[codeGPLX]": gplx,
                        "fields[birthDate]": dob,
                        "fields[birthDateType2]": "",
                        "captcha_code": cap_code.lower().strip(),
                        "securityToken": security_token,
                        "submitFormId": "8",
                        "moduleId": "8"
                    }
                    req_headers = headers.copy()
                    req_headers["Content-Type"] = "application/x-www-form-urlencoded; charset=UTF-8"
                    req_headers["Accept"] = "application/json, text/plain, */*"
                    req_headers["Origin"] = "https://gplx.csgt.bocongan.gov.vn"
                    req_headers["Referer"] = "https://gplx.csgt.bocongan.gov.vn/"

                    api_url = "https://gplx.csgt.bocongan.gov.vn/api/Project/GPLX/ApiSearchGPLX/sendRequest?site=2005782"
                    res = await client.post(api_url, data=payload, headers=req_headers)
                    result = await parse_response_to_data(res.text)

                    if not result.get("is_captcha_error"):
                        success_result = result
                        break
                    await asyncio.sleep(1)

                if success_result:
                    status = success_result.get("status")
                    await asyncio.to_thread(db.save_verification, gplx, dob, status, success_result.get("data") if status == "success" else None)
                    preload_state["done"] += 1
                    logger.info(f"[Pre-load] {gplx} → {status}")
                else:
                    preload_state["failed"] += 1
                    preload_state["errors"].append(f"{gplx}: Sai captcha liên tục")
                    if len(preload_state["errors"]) > 100: preload_state["errors"].pop(0)
                    logger.warning(f"[Pre-load] {gplx} → Thất bại (captcha)")

        except Exception as e:
            preload_state["failed"] += 1
            preload_state["errors"].append(f"{gplx}: {str(e)}")
            if len(preload_state["errors"]) > 100: preload_state["errors"].pop(0)
            logger.error(f"[Pre-load] {gplx} → Lỗi: {e}")

        await asyncio.sleep(delay)

    preload_state["running"] = False
    preload_state["current"] = ""
    logger.info(f"Pre-load hoàn tất: done={preload_state['done']}, skipped={preload_state['skipped']}, failed={preload_state['failed']}")

@app.post("/api/upload-gplx")
async def upload_gplx_image(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        
        # Fix #5: ocr.classification là synchronous → chạy trong thread pool
        text = await asyncio.to_thread(ocr.classification, contents)
        
        gplx_match = re.search(r'\b\d{12}\b', text)
        gplx_number = gplx_match.group(0) if gplx_match else ""
        
        dob_match = re.search(r'\b\d{2}[/.-]\d{2}[/.-]\d{4}\b', text)
        dob = dob_match.group(0).replace('-', '/').replace('.', '/') if dob_match else ""
        
        return JSONResponse({"status": "success", "gplx": gplx_number, "dob": dob, "raw_text": text})
    except Exception as e:
        return JSONResponse({"status": "error", "message": str(e)})

@app.post("/api/upload-bulk")
async def upload_bulk(file: UploadFile = File(...)):
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=io.BytesIO(contents))
        ws = wb.active
        
        records = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                continue
            if not row[0] or not row[1]:
                continue
                
            gplx = str(row[0]).strip()
            
            raw_dob = row[1]
            if isinstance(raw_dob, datetime):
                dob = raw_dob.strftime("%d/%m/%Y")
            else:
                dob = str(raw_dob).strip()
                
            loai = str(row[2]).strip().upper() if len(row) > 2 and row[2] else "PET"
            
            records.append({
                "gplx": gplx,
                "dob": dob,
                "loai_bang": "PET" if "PET" in loai or "2" in loai else "OLD"
            })
            
        return JSONResponse({"status": "success", "data": records})
    except Exception as e:
        return JSONResponse({"status": "error", "message": f"Không thể đọc file: {str(e)}"})

# --- WEBSOCKET FLOW ---

async def parse_response_to_data(raw_response: str):
    """Phân tích raw_response (JSON hoặc text) thành dict {status, message, data, is_captcha_error}"""
    if raw_response.strip() == "BotDetect":
        return {"status": "captcha_error", "message": "Sai mã captcha (BotDetect)", "is_captcha_error": True}
        
    try:
        data = json.loads(raw_response)
        if data:
            if isinstance(data, dict):
                if 'data' in data and data['data']:
                    return {"status": "success", "data": data['data'], "is_captcha_error": False}
                
                msg = data.get('message', '').lower()
                if 'không tìm thấy' in msg or 'not found' in msg:
                    return {"status": "not_found", "message": msg, "is_captcha_error": False}
                
                if 'captcha' in msg or 'mã bảo mật' in msg:
                    return {"status": "captcha_error", "message": "Sai Captcha", "is_captcha_error": True}
                
                if data:
                    return {"status": "success", "data": data, "is_captcha_error": False}
            elif isinstance(data, list):
                if data:
                    return {"status": "success", "data": data, "is_captcha_error": False}
                else:
                    return {"status": "not_found", "message": "Không tìm thấy dữ liệu", "is_captcha_error": False}
            else:
                return {"status": "success", "data": data, "is_captcha_error": False}
        return {"status": "not_found", "message": "Không tìm thấy dữ liệu", "is_captcha_error": False}
    except json.JSONDecodeError:
        if 'không tìm thấy' in raw_response.lower():
            return {"status": "not_found", "message": "Không tìm thấy thông tin", "is_captcha_error": False}
        elif 'thành công' in raw_response.lower():
            return {"status": "success", "data": raw_response, "is_captcha_error": False}
        elif len(raw_response) > 10:
            return {"status": "success", "data": raw_response, "is_captcha_error": False}
        return {"status": "unknown", "message": "Lỗi không xác định: " + raw_response[:50], "is_captcha_error": False}

@app.websocket("/ws/verify")
async def verify_gplx_ws(websocket: WebSocket):
    await websocket.accept()
    
    # 1. Nhận thông tin ban đầu
    try:
        init_data = await websocket.receive_json()
    except WebSocketDisconnect:
        return
        
    gplx = init_data.get("gplx", "").replace(" ", "")
    dob = init_data.get("dob", "").replace(" ", "")
    loai_bang = init_data.get("loai_bang", "PET")
    client_ip = websocket.client.host
    
    logger.info(f"[{client_ip}] Tra cứu: GPLX={gplx}, DOB={dob}, Loại={loai_bang}")
    
    # Rate Limit
    is_allowed, rl_msg = check_rate_limit(client_ip)
    if not is_allowed:
        await websocket.send_json({"type": "error", "message": rl_msg})
        await websocket.close()
        return

    # Fix #4: db calls là synchronous → chạy trong thread pool để không block event loop
    is_verified = await asyncio.to_thread(db.is_gplx_verified, gplx)
    if is_verified:
        stored_dob = await asyncio.to_thread(db.get_stored_dob, gplx)
        if stored_dob and stored_dob == dob:
            verified_data = await asyncio.to_thread(db.get_verified_data, gplx)
            logger.info(f"[{client_ip}] Kết quả từ DB: GPLX={gplx} - Thành công")
            await websocket.send_json({
                "type": "success", 
                "source": "database",
                "message": "Giấy phép này đã được xác thực trước đó.",
                "data": verified_data
            })
            await websocket.close()
            return
        elif stored_dob and stored_dob != dob:
            logger.warning(f"[{client_ip}] Sai ngày sinh cho GPLX={gplx}")
            await websocket.send_json({
                "type": "error",
                "source": "database", 
                "message": "Ngày sinh không khớp với thông tin đã xác thực trước đó."
            })
            await websocket.close()
            return

    await websocket.send_json({"type": "status", "message": "Đang kết nối hệ thống CSGT..."})
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
        "Accept-Language": "vi-VN,vi;q=0.9,en-US;q=0.8,en;q=0.7",
    }
    
    # Fix #3: Semaphore — tối đa 20 request đồng thời tới CSGT
    async with CSGT_SEMAPHORE:
        async with httpx.AsyncClient(verify=False, timeout=20.0) as client:
            try:
                # === AUTOMATIC & MANUAL CAPTCHA LOOP ===
                auto_attempts = 0
                max_auto = 2
                success_result = None
                manual_mode = False
                
                while not success_result:
                    # 1. Tải trang chủ để lấy securityToken & Cookies MỚI cho MỖI lần thử
                    resp = await client.get("https://gplx.csgt.bocongan.gov.vn/", headers=headers)
                    soup = BeautifulSoup(resp.text, 'html.parser')
                    sec_token_input = soup.find('input', {'name': 'securityToken'})
                    if not sec_token_input:
                        await websocket.send_json({"type": "error", "message": "Không thể kết nối với hệ thống CSGT (Thiếu token)."})
                        return
                    security_token = sec_token_input.get('value')
                    
                    choose_gplx = "1" if loai_bang == "OLD" else "2"
                    
                    captcha_img = soup.select_one(".img-cap-mobile img")
                    if not captcha_img:
                        await websocket.send_json({"type": "error", "message": "Không tìm thấy mã Captcha từ CSGT."})
                        return
                    captcha_url = "https://gplx.csgt.bocongan.gov.vn" + captcha_img.get('src')
                    if '?' in captcha_url:
                        captcha_url = re.sub(r't=\d+', f't={int(time.time()*1000)}', captcha_url)
                    
                    # Cần header chuẩn để tải ảnh tránh BotDetect
                    img_headers = headers.copy()
                    img_headers["Accept"] = "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8"
                    img_headers["Referer"] = "https://gplx.csgt.bocongan.gov.vn/"
                    
                    cap_resp = await client.get(captcha_url, headers=img_headers)
                    if cap_resp.status_code != 200:
                        await websocket.send_json({"type": "error", "message": "Không tải được mã Captcha."})
                        return
                    image_bytes = cap_resp.content

                    async def submit_form(cap_code: str):
                        payload_data = {
                            "type": "",
                            "fields[formTypeId]": "565f96637f8b9af6558b4567",
                            "fields[chooseGPLX]": str(choose_gplx),
                            "fields[codeGPLX]": str(gplx),
                            "fields[birthDate]": str(dob),
                            "fields[birthDateType2]": "",
                            "captcha_code": str(cap_code).lower().strip(),
                            "securityToken": str(security_token),
                            "submitFormId": "8",
                            "moduleId": "8"
                        }
                        
                        api_url = "https://gplx.csgt.bocongan.gov.vn/api/Project/GPLX/ApiSearchGPLX/sendRequest?site=2005782"
                        
                        req_headers = headers.copy()
                        req_headers["Content-Type"] = "application/x-www-form-urlencoded; charset=UTF-8"
                        req_headers["Accept"] = "application/json, text/plain, */*"
                        req_headers["Origin"] = "https://gplx.csgt.bocongan.gov.vn"
                        req_headers["Referer"] = "https://gplx.csgt.bocongan.gov.vn/"
                        
                        res = await client.post(api_url, data=payload_data, headers=req_headers)
                        return await parse_response_to_data(res.text)

                    if not manual_mode:
                        auto_attempts += 1
                        await websocket.send_json({"type": "status", "message": f"Đang tự động đọc Captcha (Lần {auto_attempts}/{max_auto})..."})
                        
                        captcha_text = await asyncio.to_thread(ocr.classification, image_bytes)
                        
                        if not captcha_text or not (4 <= len(captcha_text) <= 8):
                            await websocket.send_json({"type": "status", "message": "OCR không đọc được, tải captcha mới..."})
                            if auto_attempts >= max_auto:
                                manual_mode = True
                            continue
                            
                        await websocket.send_json({"type": "status", "message": f"Thử Captcha: {captcha_text}..."})
                        result = await submit_form(captcha_text)
                        
                        if result.get("is_captcha_error"):
                            await websocket.send_json({"type": "status", "message": "Sai Captcha tự động."})
                            await asyncio.sleep(0.5)
                            if auto_attempts >= max_auto:
                                manual_mode = True
                            continue
                        else:
                            success_result = result
                            break
                    else:
                        # Manual Mode
                        b64_img = base64.b64encode(image_bytes).decode('utf-8')
                        await websocket.send_json({
                            "type": "require_manual_captcha",
                            "image_base64": b64_img,
                            "message": "Vui lòng nhập mã Captcha bên dưới."
                        })
                        
                        try:
                            user_resp = await websocket.receive_json()
                            manual_cap = user_resp.get("captcha_code", "")
                            
                            await websocket.send_json({"type": "status", "message": "Đang gửi yêu cầu tra cứu..."})
                            result = await submit_form(manual_cap)
                            
                            if result.get("is_captcha_error"):
                                await websocket.send_json({"type": "status", "message": "Bạn đã nhập sai mã Captcha. Đang tải mã mới..."})
                                await asyncio.sleep(0.5)
                                continue # Sẽ quay lại đầu vòng lặp while để lấy captcha mới
                            else:
                                success_result = result
                                break
                                
                        except WebSocketDisconnect:
                            return

                # === XỬ LÝ KẾT QUẢ ===
                if success_result:
                    status = success_result.get("status")
                    if status == "success":
                        await asyncio.to_thread(db.save_verification, gplx, dob, "success", success_result.get("data"))
                        logger.info(f"[{client_ip}] Xác thực thành công: GPLX={gplx}")
                        await websocket.send_json({
                            "type": "success",
                            "source": "live",
                            "message": "Tra cứu thành công!",
                            "data": success_result.get("data")
                        })
                    elif status == "not_found":
                        await asyncio.to_thread(db.save_verification, gplx, dob, "not_found", None)
                        logger.info(f"[{client_ip}] Không tìm thấy: GPLX={gplx}")
                        await websocket.send_json({
                            "type": "error",
                            "source": "live",
                            "message": "Không tìm thấy thông tin trên hệ thống (Cục CSGT)."
                        })
                    else:
                        await asyncio.to_thread(db.save_verification, gplx, dob, "error", None)
                        logger.error(f"[{client_ip}] Lỗi: GPLX={gplx} - {success_result.get('message', '')}")
                        await websocket.send_json({
                            "type": "error",
                            "source": "live",
                            "message": success_result.get("message", "Lỗi không xác định.")
                        })
            
            except Exception as e:
                import traceback
                traceback.print_exc()
                await websocket.send_json({"type": "error", "message": f"Lỗi hệ thống: {str(e)}"})
                
    try:
        await websocket.close()
    except Exception:
        pass

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001, loop="asyncio")
